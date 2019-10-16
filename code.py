# Idea jest taka: all_links to lista słowników, w których są linki do OLX.
# Pętla for na podstawie linków zawartych w pojedyncznym słowniku
# wykonuje to co jest w pętli, czyli - pobranie danych z OLX i ich zapis do Excel

def pobierzListeCen():
    import requests
    from bs4 import BeautifulSoup
    # START EXCEL
    import openpyxl, datetime
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arkusz1"
    time = (datetime.datetime.now().strftime("%Y_%m_%d"))
    #FINISH EXCEL

    all_links = [{"url1": "https://www.olx.pl/nieruchomosci/stancje-pokoje/lodz/",
                  "url2": "https://www.olx.pl/nieruchomosci/stancje-pokoje/lodz/?page={}",
                  "title":"pokoje do wynajęcia"}
                ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=one",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=one&page={}",
                 "title": "kawalerki do wynajęcia"}
                 ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=two",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=two&page={}",
                  "title": "2-pokojowe mieszkania do wynajecia"}
                 ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=three",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=three&page={}",
                  "title": "3-pokojowe mieszkania do wynajecia"}
                 ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=four",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=four&page={}",
                  "title": "4 i więcej pokojowe mieszkania do wynajęcia"}
                 ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=one",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=one&page={}",
                  "title": "1 pokojowe mieszkania na sprzedaż"}
                 ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=two",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=two&page={}",
                  "title": "2 pokojowe mieszkania na sprzedaż"}
                 ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=three",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=three&page={}",
                  "title": "3 pokojowe mieszkania na sprzedaż"}
                 ,
                 {"url1": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=four",
                  "url2": "https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/lodz/"
                          "?search%5Bfilter_enum_rooms%5D%5B0%5D=four&page={}",
                  "title": "4 i wiecej pokojowe mieszkania na sprzedaż"}

                ] # do każdej głównej strony są dwa linki.
                # Pierwszy to tylko pierwsza podstrona, a drugi to druga i wszystkie następne
    lista_cen = []
    alphapet = ("ABCDEFGHIJKLMNOPQRSTUVWXYZ") #Wykorzystywany do kolumn excela
    k = 0 #Odpowiada za to w której kolumnie w Excelu jest zapis

    for links in all_links:
        lista_cen.clear()
        r = requests.get(links.get("url1"))
        print(r.status_code == requests.codes.ok)
        soup = BeautifulSoup(r.text, "lxml")

        for i, image in enumerate(soup.find_all("p", {"class": "price"})):
            if i > 4:
                lista_cen.append(float(image.text.replace("zł", "").replace(" ", "")
                                     .replace("\n", "").replace(",", ".")))

        n = 2

        while True:
            url2 = links.get("url2").format(n)
            print(url2)
            r = requests.get(url2)
            print(r.status_code == requests.codes.ok)
            print(n)
            if url2 != r.url: break
            soup = BeautifulSoup(r.text, "lxml")
            n += 1
            for i, image in enumerate(soup.find_all("p", {"class": "price"})):
                if i > 4:
                    lista_cen.append(float(image.text.replace("zł", "").replace(" ", "")
                                         .replace("\n", "").replace(",", ".")))
        #Start Excel
        for i, value in enumerate(lista_cen):
            ws["{}1".format(alphapet[k])] = links.get("title")
            ws["{}{}".format(alphapet[k], i + 2)] = value
        wb.save("ceny_kawalerek_{}.xlsx".format(time))

        k +=1
        #Finish Excel
pobierzListeCen()